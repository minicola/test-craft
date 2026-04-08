# tests/test_excel_export.py
import io
from generators.excel_export import export_to_excel
from openpyxl import load_workbook


def test_export_basic():
    cases = [
        {
            "id": "TC-001",
            "module": "登录模块",
            "title": "正常登录",
            "priority": "P0",
            "precondition": "用户已注册",
            "steps": ["打开登录页", "输入用户名", "输入密码", "点击登录"],
            "expected": "跳转到首页",
            "source": "screenshot",
        }
    ]
    output = export_to_excel(cases)
    assert isinstance(output, io.BytesIO)

    # 验证 Excel 内容
    wb = load_workbook(output)
    ws = wb.active
    assert ws.cell(row=1, column=1).value == "用例编号"
    assert ws.cell(row=2, column=1).value == "TC-001"
    assert ws.cell(row=2, column=3).value == "正常登录"
    assert "1. 打开登录页" in ws.cell(row=2, column=6).value
    assert ws.cell(row=2, column=8).value == "截图推导"


def test_export_empty():
    output = export_to_excel([])
    assert isinstance(output, io.BytesIO)
    wb = load_workbook(output)
    ws = wb.active
    assert ws.cell(row=1, column=1).value == "用例编号"
    assert ws.cell(row=2, column=1).value is None
