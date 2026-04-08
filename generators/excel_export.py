import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def export_to_excel(test_cases: list[dict]) -> io.BytesIO:
    """将测试用例导出为 Excel 文件（BytesIO 对象）。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"

    # 表头
    headers = ["用例编号", "所属模块", "用例标题", "优先级", "前置条件", "操作步骤", "预期结果", "用例来源"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # 数据行
    source_map = {"screenshot": "截图推导", "code": "代码补充", "enhanced": "历史增强",
                  "xmind": "XMind导入", "excel": "Excel导入", "ai_generated": "AI生成"}

    for row_idx, case in enumerate(test_cases, 2):
        steps_text = "\n".join(f"{i+1}. {s}" for i, s in enumerate(case.get("steps", [])))
        source = source_map.get(case.get("source", ""), case.get("source", ""))

        values = [
            case.get("id", ""),
            case.get("module", ""),
            case.get("title", ""),
            case.get("priority", "P2"),
            case.get("precondition", ""),
            steps_text,
            case.get("expected", ""),
            source,
        ]

        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin_border

    # 设置列宽
    col_widths = [12, 15, 30, 8, 20, 45, 30, 12]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # 冻结首行
    ws.freeze_panes = "A2"

    # 添加筛选器
    ws.auto_filter.ref = f"A1:H{len(test_cases) + 1}"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
