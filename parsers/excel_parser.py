from openpyxl import load_workbook

# 常见列名映射
COLUMN_MAPPINGS = {
    "module": ["模块", "所属模块", "功能模块", "module"],
    "title": ["用例名称", "用例标题", "测试用例", "标题", "title", "name"],
    "precondition": ["前置条件", "前提条件", "precondition"],
    "steps": ["操作步骤", "测试步骤", "步骤", "steps"],
    "expected": ["预期结果", "期望结果", "expected"],
    "priority": ["优先级", "级别", "priority"],
}


def parse_excel(file_path: str) -> list[dict]:
    """解析 Excel 文件，提取测试用例。"""
    wb = load_workbook(file_path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header = [str(cell).strip() if cell else "" for cell in rows[0]]
    col_map = _match_columns(header)

    cases = []
    for row in rows[1:]:
        row_data = list(row)
        if all(cell is None for cell in row_data):
            continue

        steps_raw = _get_cell(row_data, col_map.get("steps"))
        steps = _parse_steps(steps_raw) if steps_raw else []

        cases.append({
            "id": "",
            "module": _get_cell(row_data, col_map.get("module")) or "未分类",
            "title": _get_cell(row_data, col_map.get("title")) or "",
            "precondition": _get_cell(row_data, col_map.get("precondition")) or "",
            "steps": steps,
            "expected": _get_cell(row_data, col_map.get("expected")) or "",
            "priority": _normalize_priority(_get_cell(row_data, col_map.get("priority"))),
            "source": "excel",
        })
    wb.close()
    return [c for c in cases if c["title"]]


def _match_columns(header: list[str]) -> dict[str, int]:
    """将表头列名匹配到标准字段。"""
    col_map = {}
    for field, aliases in COLUMN_MAPPINGS.items():
        for i, col_name in enumerate(header):
            if col_name.lower() in [a.lower() for a in aliases]:
                col_map[field] = i
                break
    return col_map


def _get_cell(row: list, index) -> str:
    if index is None or index >= len(row):
        return ""
    val = row[index]
    return str(val).strip() if val is not None else ""


def _parse_steps(text: str) -> list[str]:
    """解析步骤文本，支持换行和编号格式。"""
    import re
    lines = text.replace("\r\n", "\n").split("\n")
    steps = []
    for line in lines:
        line = line.strip()
        if not line:
            continue
        # 去掉开头的编号（如 "1." "1、" "1)" "步骤1："）
        line = re.sub(r"^[\d]+[.、)）:：]\s*", "", line)
        if line:
            steps.append(line)
    return steps


def _normalize_priority(val: str) -> str:
    if not val:
        return "P2"
    val = val.upper().strip()
    if val in ("P0", "P1", "P2", "P3"):
        return val
    if "高" in val or "HIGH" in val:
        return "P0"
    if "中" in val or "MEDIUM" in val:
        return "P1"
    if "低" in val or "LOW" in val:
        return "P3"
    return "P2"
